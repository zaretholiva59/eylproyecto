# projects/services/excel_reports/base_reporter.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
import os

class BaseExcelReporter:
    """
    Clase base para generación de reportes Excel
    """
    
    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.styles = self._define_styles()
    
    def _define_styles(self):
        """Definir estilos base para reportes"""
        return {
            'header': Font(name='Arial', size=12, bold=True, color='FFFFFF'),
            'header_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'title': Font(name='Arial', size=14, bold=True),
            'data': Font(name='Arial', size=10),
            'currency': Font(name='Arial', size=10),
            'center_align': Alignment(horizontal='center', vertical='center')
        }
    
    def create_header(self, title, columns):
        """Crear encabezado del reporte"""
        # Título
        self.sheet.merge_cells('A1:{}1'.format(chr(64 + len(columns))))
        self.sheet['A1'] = title
        self.sheet['A1'].font = self.styles['title']
        self.sheet['A1'].alignment = self.styles['center_align']
        
        # Encabezados de columnas
        for idx, column in enumerate(columns, start=1):
            cell = self.sheet.cell(row=2, column=idx)
            cell.value = column
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center_align']
    
    def add_data_row(self, row_data, row_number):
        """Agregar fila de datos"""
        for idx, value in enumerate(row_data, start=1):
            cell = self.sheet.cell(row=row_number, column=idx)
            cell.value = value
            cell.font = self.styles['data']
    
    def auto_adjust_columns(self):
        """Ajustar automáticamente el ancho de columnas"""
        for column in self.sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            self.sheet.column_dimensions[column_letter].width = adjusted_width
    
    def save_to_response(self, filename):
        """Guardar workbook como respuesta HTTP"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        self.workbook.save(response)
        return response
    
    def save_to_file(self, filepath):
        """Guardar workbook en archivo"""
        self.workbook.save(filepath)
        return filepath
